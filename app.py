from pathlib import Path
import shutil
from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory
from dotenv import load_dotenv
import cv2
import os
import PIL.Image
from PIL import Image
import textwrap
from typing import List
import google.generativeai as genai
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook

# Google API key
load_dotenv()
GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY')
genai.configure(api_key=GOOGLE_API_KEY)
vision_model = genai.GenerativeModel(model_name="gemini-1.5-pro")
text_model = genai.GenerativeModel('gemini-1.5-flash')
roll_model = genai.GenerativeModel('gemini-1.5-flash')
model = YOLO("model/rectangle_yolo_model.pt")
model_path = YOLO("model/rectangle_yolo_model.pt")

app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads'
GENERATED_FOLDER = 'static/generated'
IMAGE_FOLDER = os.path.join(os.getcwd(), 'confirm_images')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['GENERATED_FOLDER'] = GENERATED_FOLDER

# Ensure the upload and generated folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(GENERATED_FOLDER, exist_ok=True)
app.secret_key = "your_secret_key"
app.config['UPLOAD_FOLDER_OMR'] = 'uploads'
app.config['UPLOAD_FOLDER_TEXT'] = 'uploades'

# Ensure upload folders exist
os.makedirs(app.config['UPLOAD_FOLDER_OMR'], exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER_TEXT'], exist_ok=True)

def to_markdown(text):
    text = text.replace('•', '  *')
    return textwrap.indent(text, '> ', predicate=lambda _: True)

def get_label(image_path, model_path):
    folder_path = 'predict'

    # Check if the folder exists
    if os.path.exists(folder_path):
        try:
            shutil.rmtree(folder_path)
            print(f"Folder '{folder_path}' and all its contents have been deleted.")
        except OSError as e:
            print(f"Error: {e.strerror}")
    else:
        print(f"Folder '{folder_path}' does not exist.")

    output_folder = 'predict'

    # Run prediction and save output
    results = model.predict(image_path, save=True, save_txt=True, project=output_folder, name='results', exist_ok=True)

    # Path to the folder where YOLO saves the image and label
    saved_image_folder = Path(f"{output_folder}/results")
    saved_label_folder = saved_image_folder / 'labels'

    # Find the label file that YOLO saved
    saved_labels = list(saved_label_folder.glob('*.txt'))  # Get all .txt label files in the labels folder

    # Ensure there is at least one label file
    if saved_labels:
        label_file_path = saved_labels[0]  # Assuming there's only one label file
        # Read the content of the label file into a string variable
        with open(label_file_path, 'r') as file:
            label_data = file.read()

        # Print or use the label data as needed
        print("Label Data:", label_data)
    else:
        print("No label files found.")
    return label_data


def crop_left_strip(image):
    crop = 30
    height, width = image.shape[:2]
    cropped_image = image[:, crop:]
    return cropped_image


def detect_filled_bubbles(roi):
    # Convert to grayscale
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)

    # Apply Gaussian blur
    blurred = cv2.GaussianBlur(gray, (3, 3), 0)

    # Apply thresholding
    _, binary = cv2.threshold(blurred, 90, 255, cv2.THRESH_BINARY_INV)

    # Find contours
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    detected_answers = []

    # Define bubble detection parameters
    min_area = 40  # Minimum area of a bubble contour
    max_area = 500  # Maximum area of a bubble contour
    fill_threshold = 0.3  # Percentage of area that needs to be filled to consider it marked

    # Flag to check if any bubble meets the criteria
    any_bubble_detected = False

    for contour in contours:
        area = cv2.contourArea(contour)
        if min_area < area < max_area:
            x, y, w, h = cv2.boundingRect(contour)
            bubble_roi = binary[y:y + h, x:x + w]
            filled_area = cv2.countNonZero(bubble_roi)
            if filled_area / (w * h) > fill_threshold:
                # Approximate the center of the bubble to determine its position
                cx, cy = x + w // 2, y + h // 2
                detected_answers.append((cx, cy))
                # Draw circle around detected bubble
                cv2.circle(roi, (x + w // 2, y + h // 2), 5, (0, 255, 0), 2)
                any_bubble_detected = True

    # If no bubble was detected, append '0'
    if not any_bubble_detected:
        detected_answers.append((0, 0))  # Append (0, 0) as a placeholder

    return detected_answers


# Convert the string to a 2D list
def convert_to_2d_list(data_str):
    # Split the input string into lines
    lines = data_str.strip().split('\n')
    # Split each line into a list of values and convert to float
    data_list = [list(map(float, line.split())) for line in lines]
    return data_list


# Convert the data and print the result
def final_answers(image_path, data_str):
    ## Saving the detected images in a new folder
    save_folder = "confirm_images"
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    image = cv2.imread(image_path)

    # Get image dimensions
    height, width, _ = image.shape

    # Labels (example)
    labels = convert_to_2d_list(data_str)

    # Filter class 0 boxes and sort by x_center to get them in horizontal order
    boxes = sorted([label for label in labels if label[0] == 0], key=lambda x: x[1])

    target_width = 150
    target_height = 740

    # Define option ranges
    option_ranges = {
        'A': (1, 23),
        'B': (25, 45),
        'C': (48, 67),
        'D': (72, 92),
        'E': (93, 115)
    }

    # Traverse through each box and process them
    detected_options = []
    for idx, label in enumerate(boxes):
        # Unpack the label
        class_id, center_x, center_y, w, h = label

        # Convert normalized coordinates to pixel values
        x_center = int(center_x * width)
        y_center = int(center_y * height)
        box_width = int(w * width)
        box_height = int(h * height)

        # Calculate the top-left and bottom-right corners of the bounding box
        x1 = int(x_center - box_width / 2)
        y1 = int(y_center - box_height / 2)
        x2 = int(x_center + box_width / 2)
        y2 = int(y_center + box_height / 2)

        # Extract the region of interest (ROI)
        roi = image[y1:y2, x1:x2].copy()
        roi_resize = cv2.resize(roi, (target_width, target_height))
        roi = crop_left_strip(roi_resize)

        # Calculate the height of each section
        section_height = roi.shape[0] / 50.0

        # Draw horizontal lines to divide the image into 50 parts
        for j in range(1, 50):
            y_line = int(j * section_height)
            cv2.line(roi, (0, y_line), (roi.shape[1], y_line), (0, 255, 0), 1)

        # Ensure the last line is drawn at the bottom
        cv2.line(roi, (0, roi.shape[0] - 1), (roi.shape[1], roi.shape[0] - 1), (0, 255, 0), 1)

        # Iterate through each section of the ROI
        for j in range(50):
            y_start = int(j * section_height)
            y_end = int((j + 1) * section_height)

            # Make sure the last section reaches the bottom of the image
            if j == 49:
                y_end = roi.shape[0]

            section = roi[y_start:y_end, :]

            # Detect filled bubbles within each section
            detected_answers = detect_filled_bubbles(section)

            # Map detected x-coordinates to options based on predefined ranges
            for cx, cy in detected_answers:
                for option, (min_x, max_x) in option_ranges.items():
                    if min_x <= cx < max_x:
                        detected_options.append(option)
                        break
                else:
                    # If no option matches, append '0'
                    detected_options.append('0')

        # Show the processed ROI with detected bubbles
        save_path = os.path.join(save_folder, f"rect_{idx + 1}.jpg")
        cv2.imwrite(save_path, roi)

    # Print or save the detected options for each box
    print(len(detected_options))

    return detected_options


def find_rollnumber(image_path: str, label_data: str) -> None:
    response = "Roll No."
    # Load the image
    image = cv2.imread(image_path)

    # Get image dimensions
    height, width, _ = image.shape

    # Convert the label string to a 2D list
    def convert_to_2d_list(data_str: str) -> List[List[float]]:
        lines = data_str.strip().split('\n')
        data_list = [list(map(float, line.split())) for line in lines]
        return data_list

    labels = convert_to_2d_list(label_data)

    # Filter the boxes with class 1 and save ROI
    for label in labels:
        class_id, center_x, center_y, w, h = label
        if class_id == 1:
            x_center = int(center_x * width)
            y_center = int(center_y * height)
            box_width = int(w * width)
            box_height = int(h * height)

            x1 = int(x_center - box_width / 2)
            y1 = int(y_center - box_height / 2)
            x2 = int(x_center + box_width / 2)
            y2 = int(y_center + box_width / 2)

            # Save the ROI as "roll.jpg"
            roi = image[y1:y2, x1:x2]
            cv2.imwrite("roll1.jpg", roi)
            prompt = """
            Task: Extract the 9-digit roll number from the provided image. The digits are enclosed within 9 separate boxes, so the following constraints must be strictly followed to ensure maximum accuracy.

            Accuracy Requirement:
            You must extract all 9 digits from the image with 100% consistency. The output must not change if the image is processed multiple times.
            If possible, store or cache the result of the extraction so that the output remains identical across multiple requests for the same image.
            Digit Recognition:
            Each of the 9 boxes contains exactly one digit. Ensure that you check every box to extract exactly 9 digits.
            Carefully recognize each digit inside its respective box without confusing the boundaries of the box with the digit.
            Special care must be taken to distinguish between commonly confused digits such as 0 and 6. Use boundary detection techniques to ensure digits are not misclassified due to poor alignment with the box.
            No digit should be skipped or misclassified. Ensure that the digit inside every box is read and extracted correctly.
            Additional Techniques:
            Use advanced methods like image pre-processing, boundary isolation, or digit enhancement techniques to ensure that the extraction process is accurate and consistent.
            Implement digit-box separation algorithms to help prevent misinterpretation of digits due to overlapping boundaries or poor contrast.
            Critical Review:
            After extraction, perform an internal validation pass to ensure that exactly 9 digits are extracted, and that they are correct based on the image input.
            Key Considerations:
            Consistency: The same image must always yield the same result, stored if necessary to avoid variability.
            Digit Separation: Ensure proper distinction between digits and box boundaries, especially when dealing with ambiguous digits like "0" and "6".
            No Missing Digits: Ensure that all 9 digits are extracted from their corresponding boxes.
            Explanation:
            Emphasis on 9 boxes and 9 digits ensures that each digit is accounted for and no box is skipped.
            Consistency and caching instructions help mitigate the variability of outputs for repeated image processing.
            Boundary detection and digit enhancement are reiterated to avoid common misclassification errors, particularly between "0" and "6".
            Internal validation guarantees that after extraction, no digit is missed or incorrectly processed, ensuring all 9 boxes are captured properly.
            **Do not write anything from your side just provide the detected roll number as output**
            """
            roll_image = PIL.Image.open('roll1.jpg')
            response = roll_model.generate_content([prompt, roll_image])
            response = response.text
            print(response)

    return response
def chq(detected_list, questions):
    words = list(questions)
    words = [item.lower() for item in words]
    detected_list = [item.lower() for item in detected_list]

    wrong_ans = []
    # total_marks = sum(1 for i in range(len(words)) if detected_list[i] == words[i])
    total_marks = 0
    for i in range(len(words)):
        if words[i] == detected_list[i]:
            total_marks += 1
        else:
            wrong_ans.append(i + 1)
    total_marks = str(total_marks) + " / " + str(len(words))

    wrong_answer = "<br>".join(f"Question {i}" for i in wrong_ans)
    return total_marks, wrong_answer

def feed_record(roll_number, marks_obtained):
    # File path
    excel_file = "student_marks.xlsx"

    # Create or load the Excel sheet
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        # Add headers
        sheet["A1"] = "Roll Number"
        sheet["B1"] = "Marks Obtained"

    # Function to check if roll number already exists
    def roll_number_exists(roll_number):
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == roll_number:
                return True
        return False

    # Function to add entry
    def add_entry(roll_number, marks_obtained):
        if not roll_number_exists(roll_number):
            # Find the next available row
            next_row = sheet.max_row + 1
            # Store data
            sheet[f"A{next_row}"] = roll_number
            sheet[f"B{next_row}"] = marks_obtained
            # Save the workbook
            workbook.save(excel_file)
            print(f"Entry added: Roll Number: {roll_number}, Marks Obtained: {marks_obtained}")
        else:
            print(f"Duplicate entry avoided for Roll Number: {roll_number}.")
    add_entry(roll_number, marks_obtained)

roll_num=None
total_marks=None
wrong_answers=None

@app.route('/')
def home():
    return render_template('switch.html')

@app.route('/upload_omr', methods=['GET', 'POST'])
def upload_file():
    global total_marks
    global wrong_answers
    global roll_num
    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            actual_ans = request.form.get("question")
            labels = get_label(file_path, model_path)
            roll_num = find_rollnumber(file_path, labels)
            detected_ans = final_answers(file_path, labels)
            total_marks, wrong_answers = chq(detected_ans, actual_ans)
            return redirect(url_for('display', eval_ans = detected_ans, eval_roll = roll_num))

    return render_template('upload.html')

@app.route('/display')
def display():
    eval_ans = request.args.getlist('eval_ans')
    eval_roll = request.args.getlist('eval_roll')
    print(eval_roll)
    print(eval_ans)
    # List of image filenames
    images = [
        'rect_1.jpg',
        'rect_2.jpg',
        'rect_3.jpg',
    ]
    return render_template('display.html', images=images, eval_ans=eval_ans, eval_roll=eval_roll)

@app.route('/images/<filename>')
def get_image(filename):
    # Serve the image from the 'new' directory
    return send_from_directory(IMAGE_FOLDER, filename)

@app.route('/final_evaluation', methods=['POST'])
def final_evaluation():
    feed_record(roll_num, total_marks)
    return render_template('result.html', total_marks=total_marks, wrong_answer=wrong_answers)

@app.route('/text')
def text_page():
    return render_template('text_index.html')


@app.route('/predict', methods=['POST'])
def predict():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file:
        filepath = os.path.join(app.config['UPLOAD_FOLDER_TEXT'], file.filename)
        file.save(filepath)

        image = Image.open(filepath)
        output = "Got"
        if image is None:
            output = "Not Got"
        else:
            evaluation_prompt = '''
                You are an expert examiner with 20 years of experience in evaluating answers across various subjects. You will evaluate the provided question and answer based on the following marking scheme:
                Marking Scheme:
                Meaning and Accuracy (4 marks)
                Evaluate if the answer correctly and comprehensively addresses the question.
                Award up to 4 marks based on the accuracy and completeness of the response.
                Grammar (3 marks)
                Check for grammatical errors.
                Award 3 marks if no errors are found.
                Deduct marks for each grammatical mistake identified.
                Language Level (3 marks)
                Assess the language proficiency of the answer.
                Award 3 marks for advanced-level language, 2 marks for medium-level, and 1 mark for poor-level language.

                Output Format:
                Overall Score: Provide the total score in the format X/10.

                Evaluation Criteria:
                Be versatile and capable of evaluating answers from any subject or domain.
                Clearly indicate the points of mistakes and suggest corrections where necessary.
                Provide a detailed breakdown of the score, with proper gaps and indexing for clarity.
                Highlight the strengths and weaknesses of the answer.
                Ensure the feedback is structured in a user-friendly manner, making it easy to understand.

                Example Evaluation:
                Question: Describe the process of photosynthesis and its significance in plants.

                Answer: Photosynthesis is the process by which green plants, algae, and some bacteria convert light energy into chemical energy in the form of glucose. This process occurs in the chloroplasts of plant cells and involves several steps. In the first step, light energy is absorbed by chlorophyll and other pigments in the chloroplasts. This energy is used to split water molecules into oxygen, protons, and electrons. The oxygen is released into the atmosphere, while the protons and electrons are used to convert carbon dioxide into glucose. Photosynthesis is crucial for plants as it provides them with the energy they need for growth, reproduction, and other metabolic processes. It is also important for the environment as it produces oxygen, which is essential for the survival of most living organisms.

                Evaluation:
                Overall Score: 10/10

                Instructions for Awarding Marks
                General Guidelines:
                Consistency is Key: Once you assign marks to a specific question-answer pair, those marks cannot be changed.
                Uniformity: If the same question-answer pair is asked by different students, they should receive the same marks as initially awarded.
                Example:
                Initial Marking:

                Student A asks a question.
                You award 7 marks for the question-answer pair.
                Subsequent Marking:

                Student B asks the same question.
                Award 7 marks to Student B as well, maintaining consistency with the initial marking.
                Important Notes:
                No Retroactive Changes: Do not adjust marks for previously evaluated pairs.
                Equity: Ensure all students are evaluated fairly based on the established marks for each question-answer pair.
                Remember one thing:- Always provide overall ans on top of the generated response.
                '''
            question = request.form['question']
            response = vision_model.generate_content(["What is written in the image", image], stream=True)
            response.resolve()
            extracted_text = response.text
            prompt = f"{evaluation_prompt}, Given the content above, {question} and answer is {extracted_text}"

            evaluation_response = text_model.generate_content(prompt)
            marks = evaluation_response.candidates[0].content.parts[0].text

        return render_template('text_index.html', prediction_text=extracted_text, evaluated_marks=marks)


@app.route('/logout')
def logout():
    session.pop('email', None)
    return redirect(url_for('home'))

if __name__ == "__main__":
    app.run(debug=True)
